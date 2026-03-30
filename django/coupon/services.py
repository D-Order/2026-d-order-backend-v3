import secrets
import string
from decimal import Decimal
from django.db import transaction
from django.db.models import Count, Q
from django.shortcuts import get_object_or_404
from django.utils import timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from booth.models import *
from table.models import *
from cart.models import *
from cart.services import *

from .models import *

class CouponError(Exception):
    def __init__(self, message, error_code="COUPON_ERROR", detail=None, status_code=400):
        super().__init__(message)
        self.message = message
        self.error_code = error_code
        self.detail = detail or message
        self.status_code = status_code


def _ensure_table_usage_alive(table_usage: TableUsage):
    if table_usage.ended_at is not None:
        raise CouponError("세션이 종료되었습니다.", "TABLE_USAGE_ENDED", status_code=410)


def _ensure_cart_active(cart: Cart):
    if cart.status != Cart.Status.ACTIVE:
        raise CouponError("active 상태에서만 쿠폰을 변경할 수 있습니다.", "CART_NOT_ACTIVE", status_code=409)


def _gen_code(length: int = 6) -> str:
    alphabet = string.ascii_uppercase + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _calc_discount(subtotal: int, discount_type: str, discount_value: Decimal) -> int:
    if subtotal <= 0:
        return 0
    if discount_type == "RATE":
        amt = int(Decimal(subtotal) * discount_value)
        return max(0, min(subtotal, amt))
    amt = int(discount_value)
    return max(0, min(subtotal, amt))


@transaction.atomic
def create_coupon_and_codes(
    *,
    booth: Booth,
    name: str,
    description: str | None,
    discount_type: str,
    discount_value: Decimal,
    quantity: int,
) -> Coupon:

    if quantity < 1:
        raise CouponError("quantity는 1 이상이어야 합니다.", "VALIDATION_ERROR", status_code=400)

    coupon = Coupon.objects.create(
        booth=booth,
        name=name,
        description=description if description else None,
        discount_type=discount_type,
        discount_value=discount_value,
        quantity=quantity,
    )

    created = 0
    while created < quantity:
        code = _gen_code(6)
        try:
            CouponCode.objects.create(coupon=coupon, code=code)
            created += 1
        except Exception:
            continue

    return coupon


def list_coupons_for_booth(*, booth: Booth):
    qs = (
        Coupon.objects.filter(booth=booth)
        .annotate(used_count=Count("codes", filter=Q(codes__used_at__isnull=False)))
        .order_by("-created_at")
    )
    return qs


@transaction.atomic
def delete_coupon_if_unused(*, coupon_id: int) -> None:
    coupon = get_object_or_404(Coupon.objects.select_for_update(), id=coupon_id)

    used_exists = CouponCode.objects.filter(coupon=coupon, used_at__isnull=False).exists()
    applied_exists = CartCouponApply.objects.filter(coupon_code__coupon=coupon).exists()

    if used_exists or applied_exists:
        raise CouponError("이미 사용된(또는 적용된) 쿠폰은 삭제할 수 없습니다.", "COUPON_ALREADY_USED", status_code=409)

    coupon.delete()


@transaction.atomic
def apply_coupon_code(*, table_usage_id: int, coupon_code_str: str):
    table_usage = get_object_or_404(TableUsage, id=table_usage_id)
    _ensure_table_usage_alive(table_usage)

    cart = get_object_or_404(Cart.objects.select_for_update(), table_usage=table_usage)
    _ensure_cart_active(cart)

    code = get_object_or_404(CouponCode.objects.select_for_update(), code=coupon_code_str)

    booth_id = table_usage.table.booth_id
    if code.coupon.booth_id != booth_id:
        raise CouponError("해당 부스에서 사용할 수 없는 쿠폰입니다.", "COUPON_BOOTH_MISMATCH", status_code=400)

    if code.used_at is not None:
        raise CouponError("이미 사용된 쿠폰 코드입니다.", "COUPON_CODE_USED", status_code=400)

    if CartCouponApply.objects.filter(cart=cart, round=cart.round).exists():
        raise CouponError("이미 쿠폰이 적용되어 있습니다.", "COUPON_ALREADY_APPLIED", status_code=409)

    subtotal = recalc_cart_price(cart)
    discount_amount = _calc_discount(subtotal, code.coupon.discount_type, code.coupon.discount_value)
    total = subtotal - discount_amount

    CartCouponApply.objects.create(cart=cart, round=cart.round, coupon_code=code)

    return {
        "coupon": {
            "coupon_id": code.coupon_id,
            "coupon_code": code.code,
            "discount_type": code.coupon.discount_type,
            "discount_value": code.coupon.discount_value,
            "discount_amount": discount_amount,
        },
        "summary": {
            "subtotal": subtotal,
            "discount_total": discount_amount,
            "total": total,
        },
        "round": cart.round,
    }


@transaction.atomic
def cancel_coupon_apply(*, table_usage_id: int):
    table_usage = get_object_or_404(TableUsage, id=table_usage_id)
    _ensure_table_usage_alive(table_usage)

    cart = get_object_or_404(Cart.objects.select_for_update(), table_usage=table_usage)
    _ensure_cart_active(cart)

    applied = (
        CartCouponApply.objects.select_for_update()
        .filter(cart=cart, round=cart.round)
        .select_related("coupon_code")
        .first()
    )
    if not applied:
        raise CouponError("적용된 쿠폰이 없습니다.", "COUPON_NOT_APPLIED", status_code=400)

    applied.delete()

    subtotal = recalc_cart_price(cart)
    return {
        "summary": {"subtotal": subtotal, "discount_total": 0, "total": subtotal},
        "round": cart.round,
    }
    
def build_coupon_excel_for_booth(*, booth: Booth) -> bytes:
    coupons = (
        Coupon.objects.filter(booth=booth)
        .prefetch_related("codes")
        .order_by("-created_at", "id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "쿠폰 목록"

    ws.append([
        "쿠폰 이름",
        "쿠폰 내용",
        "할인 타입",
        "할인 값",
        "쿠폰 재고",
        "코드",
        "사용 여부",
    ])

    for coupon in coupons:
        codes = coupon.codes.all().order_by("created_at")

        if not codes.exists():
            ws.append([
                coupon.name,
                coupon.description or "",
                coupon.get_discount_type_display(),
                float(coupon.discount_value),
                coupon.quantity,
                "",
                "미사용",
            ])
            continue

        for code in codes:
            ws.append([
                coupon.name,
                coupon.description or "",
                coupon.get_discount_type_display(),
                float(coupon.discount_value),
                coupon.quantity,
                code.code,
                "사용" if code.used_at else "미사용",
            ])

    widths = [20, 30, 15, 12, 12, 18, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()

def get_coupon_detail_with_codes(*, booth: Booth, coupon_id: int, status: str = "ALL"):
    coupon = get_object_or_404(
        Coupon.objects.filter(booth=booth).annotate(
            used_count=Count("codes", filter=Q(codes__used_at__isnull=False))
        ),
        id=coupon_id,
    )

    code_qs = coupon.codes.all().order_by("created_at", "id")

    status = (status or "ALL").upper()
    if status == "USED":
        code_qs = code_qs.filter(used_at__isnull=False)
    elif status == "UNUSED":
        code_qs = code_qs.filter(used_at__isnull=True)
    elif status != "ALL":
        raise CouponError(
            "status 값이 올바르지 않습니다.",
            error_code="INVALID_STATUS",
            detail="status must be one of ALL, USED, UNUSED",
            status_code=400,
        )

    used_count = coupon.used_count or 0
    unused_count = max(0, coupon.quantity - used_count)

    if coupon.discount_type == Coupon.DiscountType.RATE:
        display_discount_value = float(coupon.discount_value) * 100
    else:
        display_discount_value = float(coupon.discount_value)

    coupon_data = {
        "coupon_id": coupon.id,
        "name": coupon.name,
        "description": coupon.description,
        "discount_type": coupon.discount_type,
        "discount_value": float(coupon.discount_value),
        "display_discount_value": display_discount_value,
        "quantity": coupon.quantity,
        "used_count": used_count,
        "unused_count": unused_count,
        "created_at": coupon.created_at,
    }

    codes_data = [
        {
            "coupon_code_id": code.id,
            "code": code.code,
            "is_used": code.used_at is not None,
            "used_at": code.used_at,
            "created_at": code.created_at,
        }
        for code in code_qs
    ]

    return {
        "coupon": coupon_data,
        "codes": codes_data,
    }