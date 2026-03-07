from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from booth.models import *
from .serializers import *
from .services import *
from .models import *
from cart.services_ws import broadcast_cart_event


def error_response(e: CouponError):
    return Response(
        {
            "status": "error",
            "code": e.status_code,
            "message": e.message,
            "data": {"error_code": e.error_code, "detail": e.detail},
        },
        status=e.status_code,
    )


def get_admin_booth(request) -> Booth:
    if not request.user.is_staff:
        raise CouponError("권한이 없습니다.", error_code="FORBIDDEN", detail="admin only", status_code=403)

    try:
        return request.user.booth
    except Booth.DoesNotExist:
        raise CouponError(
            "운영자 부스 정보를 찾을 수 없습니다.",
            error_code="BOOTH_NOT_FOUND",
            detail="user has no booth mapped",
            status_code=404,
        )


# 운영자용: 쿠폰 목록/등록
class CouponListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            booth = get_admin_booth(request)
        except CouponError as e:
            return error_response(e)

        qs = list_coupons_for_booth(booth=booth)

        data = []
        for c in qs:
            total_count = c.quantity
            remaining_count = max(0, total_count - (c.used_count or 0))
            data.append(
                {
                    "coupon_id": c.id,
                    "name": c.name,
                    "discount_type": c.discount_type,
                    "discount_value": float(c.discount_value),
                    "created_at": c.created_at,
                    "total_count": total_count,
                    "remaining_count": remaining_count,
                }
            )

        return Response({"status": "success", "code": 200, "data": data}, status=200)

    def post(self, request):
        try:
            booth = get_admin_booth(request)
        except CouponError as e:
            return error_response(e)

        serializer = CouponCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        payload = dict(serializer.validated_data)
        payload.pop("booth_id", None)

        try:
            coupon = create_coupon_and_codes(booth=booth, **payload)
        except CouponError as e:
            return error_response(e)
        except Exception as e:
            return Response(
                {"message": "쿠폰 등록 중 오류가 발생했습니다.", "data": {"error_code": "COUPON_CREATE_FAILED", "detail": str(e)}},
                status=500,
            )

        return Response(
            {
                "message": "쿠폰이 등록되었습니다",
                "data": {
                    "coupon": {
                        "coupon_id": coupon.id,
                        "booth_id": coupon.booth_id,
                        "name": coupon.name,
                        "description": coupon.description,
                        "discount_type": coupon.discount_type,
                        "discount_value": float(coupon.discount_value),
                        "quantity": coupon.quantity,
                        "created_at": coupon.created_at,
                    }
                },
            },
            status=200,
        )


# 운영자용: 쿠폰 삭제
class CouponDeleteAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, coupon_id: int):
        try:
            booth = get_admin_booth(request)
        except CouponError as e:
            return error_response(e)
        coupon = get_object_or_404(Coupon, id=coupon_id, booth=booth)

        try:
            delete_coupon_if_unused(coupon_id=coupon.id)
        except CouponError as e:
            return error_response(e)

        return Response({"status": "success", "code": 200, "message": f"쿠폰 {coupon_id}가 삭제되었습니다."}, status=200)


# 운영자용: 쿠폰 코드 엑셀 다운로드
class CouponDownloadAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, coupon_id: int):
        try:
            booth = get_admin_booth(request)
        except CouponError as e:
            return error_response(e)
        coupon = get_object_or_404(Coupon, id=coupon_id, booth=booth)
        codes = CouponCode.objects.filter(coupon=coupon).order_by("created_at")

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from django.utils import timezone

        wb = Workbook()
        ws = wb.active
        ws.title = "coupon_codes"

        ws.append(["coupon_id", "coupon_name", "code", "used", "used_at", "created_at"])
        for code in codes:
            ws.append(
                [
                    coupon.id,
                    coupon.name,
                    code.code,
                    bool(code.used_at),
                    code.used_at.isoformat() if code.used_at else None,
                    code.created_at.isoformat() if code.created_at else None,
                ]
            )

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"booth_{coupon.booth_id}_coupon_{coupon.id}_codes_{timezone.now().strftime('%Y%m%d')}.xlsx"
        resp = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


# 손님용: 쿠폰 적용/취소
class CouponApplyAPIView(APIView):
    def post(self, request):
        serializer = CouponApplySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        table_usage_id = serializer.validated_data["table_usage_id"]
        
        try:
            result = apply_coupon_code(
                table_usage_id=serializer.validated_data["table_usage_id"],
                coupon_code_str=serializer.validated_data["coupon_code"],
            )
        except CouponError as e:
            return Response({"message": e.message, "data": {"error_code": e.error_code, "detail": e.detail}}, status=e.status_code)
        
        broadcast_cart_event(
            table_usage_id=table_usage_id,
            event_type="CART_COUPON_APPLIED",
            message="쿠폰이 적용되었습니다.",
        )

        return Response({"message": "쿠폰이 적용되었습니다", "data": result}, status=200)

    def delete(self, request):
        serializer = CouponCancelSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        table_usage_id = serializer.validated_data["table_usage_id"]
        
        try:
            result = cancel_coupon_apply(table_usage_id=serializer.validated_data["table_usage_id"])
        except CouponError as e:
            return Response({"message": e.message, "data": {"error_code": e.error_code, "detail": e.detail}}, status=e.status_code)

        broadcast_cart_event(
            table_usage_id=table_usage_id,
            event_type="CART_COUPON_CANCELLED",
            message="쿠폰 적용이 취소되었습니다.",
        )
        
        return Response({"message": "쿠폰 적용이 취소되었습니다", "data": result}, status=200)